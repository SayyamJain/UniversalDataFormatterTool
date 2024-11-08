import os
import json
import csv
import xml.etree.ElementTree as ET
import openpyxl
import yaml
import tkinter as tk
from collections import OrderedDict
from tkinter import filedialog, messagebox
from tkinter import ttk


def read_json(file_path):
    try:
        with open(file_path, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        messagebox.showerror("Error", f"File {file_path} not found.")
        return None
    except json.JSONDecodeError:
        messagebox.showerror("Error", f"Error decoding JSON from {file_path}.")
        return None
    except Exception:
        messagebox.showerror(
            "Error",
            f"Error decoding JSON from {file_path}. Error message: Not a valid JSON file!",
        )
        return None


def flatten_json(data):
    flat_data = []

    def flatten(item, parent_key=""):
        flat_item = OrderedDict()
        for key, value in item.items():
            new_key = f"{parent_key}_{key}" if parent_key else key
            if isinstance(value, dict):
                flat_item.update(flatten(value, new_key))
            elif isinstance(value, list):
                for i, sub_item in enumerate(value):
                    if isinstance(sub_item, dict):
                        flat_item.update(flatten(sub_item, f"{new_key}_{i}"))
                    else:
                        flat_item[f"{new_key}_{i}"] = sub_item
            else:
                flat_item[new_key] = value
        return flat_item

    if isinstance(data, list):
        for item in data:
            flat_data.append(flatten(item))
    else:
        flat_data.append(flatten(data))

    return flat_data


def write_json(data, file_path):
    try:
        with open(file_path, "w") as f:
            json.dump(data, f, indent=4)
    except Exception as e:
        messagebox.showerror("Error", f"Error writing to {file_path}: {e}")


def read_csv(file_path):
    try:
        with open(file_path, "r") as f:
            return list(csv.DictReader(f))
    except FileNotFoundError:
        messagebox.showerror("Error", f"File {file_path} not found.")
        return None
    except csv.Error:
        messagebox.showerror("Error", f"Error reading CSV from {file_path}.")
        return None
    except Exception:
        messagebox.showerror(
            "Error",
            f"Error decoding JSON from {file_path}. Error message: Not a valid CSV file!",
        )
        return None


def write_csv(data, file_path):
    try:
        if data:
            # Use the first item's keys as the fieldnames to preserve order
            fieldnames = list(data[0].keys())

            with open(file_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
        else:
            print("Error: No data to write to CSV.")
    except Exception as e:
        print(f"Error writing to {file_path}: {e}")


def read_xml(file_path):
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        return [{elem.tag: elem.text for elem in child} for child in root]
    except FileNotFoundError:
        messagebox.showerror("Error", f"File {file_path} not found.")
        return None
    except ET.ParseError:
        messagebox.showerror("Error", f"Error parsing XML from {file_path}.")
        return None
    except Exception:
        messagebox.showerror(
            "Error",
            f"Error decoding JSON from {file_path}. Error message: Not a valid XML file!",
        )
        return None


def write_xml(data, file_path):
    try:

        def build_xml_element(item, parent):
            for key, value in item.items():
                if isinstance(value, dict):
                    child = ET.SubElement(parent, key)
                    build_xml_element(value, child)
                elif isinstance(value, list):
                    for i, sub_item in enumerate(value):
                        list_item = ET.SubElement(parent, f"{key}_{i}")
                        if isinstance(sub_item, dict):
                            build_xml_element(sub_item, list_item)
                        else:
                            list_item.text = str(sub_item)
                else:
                    child = ET.SubElement(parent, key)
                    child.text = str(value)

        root = ET.Element("root")
        if isinstance(data, list):
            for item in data:
                item_element = ET.SubElement(root, "item")
                build_xml_element(item, item_element)
        else:
            build_xml_element(data, root)

        tree = ET.ElementTree(root)
        tree.write(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Error writing to {file_path}: {e}")


def read_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        data = []
        headers = [cell.value for cell in sheet[1]]  # type: ignore
        for row in sheet.iter_rows(min_row=2, values_only=True):  # type: ignore
            data.append(dict(zip(headers, row)))
        return data
    except FileNotFoundError:
        messagebox.showerror("Error", f"File {file_path} not found.")
        return None
    except Exception as e:
        messagebox.showerror("Error", f"Error reading Excel file {file_path}: {e}")
        return None


def write_excel(data, file_path):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        if data:
            headers = list(data[0].keys())
            sheet.append(headers)
            for item in data:
                sheet.append([item.get(key, "") for key in headers])
        wb.save(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Error writing to Excel file {file_path}: {e}")


def read_yaml(file_path):
    try:
        with open(file_path, "r") as f:
            return yaml.safe_load(f)
    except FileNotFoundError:
        messagebox.showerror("Error", f"File {file_path} not found.")
        return None
    except yaml.YAMLError:
        messagebox.showerror("Error", f"Error reading YAML file {file_path}.")
        return None
    except Exception:
        messagebox.showerror(
            "Error",
            f"Error decoding JSON from {file_path}. Error message: Not a valid YAML file!",
        )
        return None


def write_yaml(data, file_path):
    try:
        # Convert OrderedDict to regular dict
        data = [dict(item) for item in data]
        with open(file_path, "w") as f:
            yaml.safe_dump(data, f)
    except Exception as e:
        messagebox.showerror("Error", f"Error writing to YAML file {file_path}: {e}")


def convert_data(input_path, output_path, input_format, output_format):
    if not os.path.isfile(input_path):
        status_label.config(text="Source file does not exist.")
        messagebox.showerror("Error", f"Source file {input_path} does not exist.")
        return

    status_label.config(text="Reading source file...")
    root.update_idletasks()

    if input_format == "json":
        data = read_json(input_path)
        data = flatten_json(data)
    elif input_format == "csv":
        data = read_csv(input_path)
    elif input_format == "xml":
        data = read_xml(input_path)
    elif input_format == "excel":
        data = read_excel(input_path)
    elif input_format == "yaml":
        data = read_yaml(input_path)
    else:
        messagebox.showerror(
            "Error",
            "Unsupported input format! Supported formats: json, csv, xml.",
        )
        return

    if data is None:
        status_label.config(text="Failed to read source file.")
        return

    status_label.config(text="Writing output file...")
    root.update_idletasks()

    # If output path is not provided, generate it based on the input file path and the output format
    if not output_path:
        output_path = os.path.splitext(input_path)[0] + "." + output_format

    if output_format == "json":
        write_json(data, output_path)
    elif output_format == "csv":
        write_csv(data, output_path)
    elif output_format == "xml":
        write_xml(data, output_path)
    elif output_format == "excel" or output_format == "xlsx":
        write_excel(data, output_path)
    elif output_format == "yaml":
        write_yaml(data, output_path)
    else:
        messagebox.showerror(
            "Error",
            "Unsupported output format! Supported formats: json, csv, xml, excel, yaml.",
        )

    status_label.config(text="Conversion completed.")
    messagebox.showinfo("Success", f"Data successfully converted to {output_path}")


def select_input_file():
    file_path = filedialog.askopenfilename()
    input_path_entry.delete(0, tk.END)
    input_path_entry.insert(0, file_path)


def select_output_file():
    # Get the selected output format
    selected_format = output_format_var.get().lower()

    if selected_format == "excel":
        selected_format = "xlsx"

    # Open the save file dialog
    file = filedialog.asksaveasfilename(defaultextension=f".{selected_format}")

    # Set the file path with the correct extension
    output_path_entry.delete(0, tk.END)
    output_path_entry.insert(0, file)


def on_browse_input():
    input_path = filedialog.askopenfilename()
    input_path_entry.delete(0, tk.END)
    input_path_entry.insert(0, input_path)

    # Automatically set the output path based on the input path
    output_path = os.path.splitext(input_path)[0]
    output_path_entry.delete(0, tk.END)
    output_path_entry.insert(0, output_path)


def on_convert():
    input_path = input_path_entry.get()
    output_path = output_path_entry.get()
    input_format = input_format_var.get().lower()
    output_format = output_format_var.get().lower()

    if output_format == "excel":
        output_format = "xlsx"

    status_label.config(text="Starting conversion...")
    root.update_idletasks()

    # Checks if output file already have some extension
    hasExtension = output_path.endswith(output_format)

    if hasExtension:
        output_path = output_path
    else:
        # Append the output format as the extension to the output path
        output_path += "." + output_format
    convert_data(input_path, output_path, input_format, output_format)


def reset_inputs():
    input_path_entry.delete(0, "end")
    output_path_entry.delete(0, "end")
    input_format_var.set("JSON")
    output_format_var.set("JSON")
    status_label.config(text="")


# Create the main window
root = tk.Tk()
root.title("Data Formatter Tool")

# Use a themed style
style = ttk.Style()
style.theme_use("clam")

# Main frame
main_frame = ttk.Frame(root, padding="10")
main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))  # type: ignore

# Configure grid to be responsive
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
main_frame.columnconfigure(0, weight=1)
main_frame.columnconfigure(1, weight=1)
main_frame.columnconfigure(2, weight=1)
main_frame.rowconfigure(0, weight=1)
main_frame.rowconfigure(1, weight=1)
main_frame.rowconfigure(2, weight=1)
main_frame.rowconfigure(3, weight=1)
main_frame.rowconfigure(4, weight=1)
main_frame.rowconfigure(5, weight=1)
main_frame.rowconfigure(6, weight=1)

# Button frame to hold Convert, Exit, and Reset buttons
button_frame = ttk.Frame(main_frame)
button_frame.grid(row=5, column=0, columnspan=3, pady=10)

# Input file selection
ttk.Label(main_frame, text="Source File:").grid(
    row=0, column=0, padx=5, pady=5, sticky=tk.E
)
input_path_entry = ttk.Entry(main_frame, width=50)
input_path_entry.grid(row=0, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
ttk.Button(main_frame, text="Browse...", command=on_browse_input).grid(
    row=0, column=2, padx=5, pady=5
)

# Input format selection
ttk.Label(main_frame, text="Format:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)
input_format_var = tk.StringVar()
input_format_combobox = ttk.Combobox(
    main_frame, textvariable=input_format_var, state="readonly", width=48
)
input_format_combobox["values"] = ["JSON", "CSV", "XML", "Excel", "YAML"]
input_format_combobox.grid(row=1, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
input_format_combobox.set("JSON")

# Output file selection
ttk.Label(main_frame, text="Destination File:").grid(
    row=2, column=0, padx=5, pady=5, sticky=tk.E
)
output_path_entry = ttk.Entry(main_frame, width=50)
output_path_entry.grid(row=2, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
ttk.Button(main_frame, text="Browse...", command=select_output_file).grid(
    row=2, column=2, padx=5, pady=5
)

# Output format selection
ttk.Label(main_frame, text="Format:").grid(row=3, column=0, padx=5, pady=5, sticky=tk.E)
output_format_var = tk.StringVar()
output_format_combobox = ttk.Combobox(
    main_frame, textvariable=output_format_var, state="readonly", width=48
)
output_format_combobox["values"] = ["JSON", "CSV", "XML", "Excel", "YAML"]
output_format_combobox.grid(row=3, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
output_format_combobox.set("JSON")

# Convert button
ttk.Button(button_frame, text="Convert", command=on_convert).grid(
    row=0, column=0, padx=10, pady=5
)

# Exit button
ttk.Button(button_frame, text="  Exit  ", command=root.destroy).grid(
    row=0, column=1, padx=10, pady=5
)

# Reset button
ttk.Button(button_frame, text="Reset", command=reset_inputs).grid(
    row=0, column=2, padx=10, pady=5
)

# Status label
status_label = ttk.Label(
    main_frame,
    text="",
    background="light gray",
    foreground="black",
    font=("Helvetica", 10, "bold"),
)
status_label.grid(row=6, column=0, columnspan=3, padx=5, pady=5)

# Start the GUI event loop
root.mainloop()
